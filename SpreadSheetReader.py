import pandas as pd # type: ignore
import requests
import requests 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

#Only focus on first two colums, which include the first and last name, respectively
require_cols = [0, 1]

#Create a dataframe using the information collected from the first and second columns
required_df = pd.read_excel('mexico commonality list copy.xlsx', usecols = require_cols, dtype= {'first name':str, 'last name':str})

#Separate the dataframe into two lists, one with the first names and the other with the last names
first_names = required_df['first name'].tolist()
last_names = required_df['last name'].tolist()

#Create a list that will hold the first and last names of each person together
check_names = []

#Go through every name in first_names, concatenate the corresponding last_name to it, and store it in check_names
x = 0
while x < len(first_names):
    check_names.append(first_names[x] + " " + last_names[x])
    x += 1

'''Using the Zendesk Sell Contacts API, this function generates a dictionary, based on the check_names list, that holds index values as keys and lists as values.
   The each list holds two values, the name of the person and an int (0 or 1).
   If the name is shortened, contains a special char, or appears on Zendesk it will be a 1. but if it isn't any of those it is a 0.
''' 
def contactsChecker(names):
    data_dict = defaultdict(list)

    #Any special characters that might be included in a name
    #Also any possible shortnames that could be included - This could be improved by using AI to seek out potential nicknames, instead of using a list
    special_chars = ['.', '-', '\'']
    shortened_names = ['Nelly', 'Anabel', 'Beto', 'Pancho', 'Paco', 'Pepe', 'Alejo', 'Lalo', 'Memo', 'Eddie', 'Frank']

    #Access to the Zendesk API - go to https://developer.zendesk.com/documentation/sales-crm/first-call/ for more info
    headers = {'Accept' : 'application/json', 'Authorization' : 'Bearer 3cf9249545d38c1d94c16457731eed492b4ead4d618819de0d633f7c3cb09346'}

    y = 0

    for name in names:
        pair = []
        pair.append(name)

        # Go to https://developer.zendesk.com/api-reference/sales-crm/resources/contacts/ for more info
        url = f'https://api.getbase.com/v2/contacts/?name={name}'
        response = requests.get(url, headers = headers)

        if any(x in name for x in special_chars) or any(x in name for x in shortened_names):
            pair.append(1)
            data_dict[y] = pair
            y += 1
            continue
        elif not response.json()['items']:
            pair.append(0)
        elif response:
            pair.append(1)
        else:
            raise Exception(f"Non-success status code: {response.status_code}")
        
        data_dict[y] = pair
        y += 1
    return data_dict
 

result_names = contactsChecker(check_names)

#assigns a color to the row the name is in, yellow or green, depending on the value associated to it
#0 = green, 1 = yellow
#this results is built into the original file given

file = 'mexico commonality list copy.xlsx'
wb = load_workbook(filename = file)
ws = wb['Sheet1']

x= 2
for key, val in result_names.items():
    if val[1] < 1:
        for cell in ws[f"{x}:{x}"]:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
    else:
        for cell in ws[f"{x}"]:
            cell.fill = PatternFill(fgColor="FFFF00", fill_type="solid")
    x += 1

wb.save(filename=file)